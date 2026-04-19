from __future__ import annotations

import os
from io import BytesIO
from pathlib import Path

from django.conf import settings
from django.http import HttpResponse

# ---------------------------------------------------------------------------
# Brand colours
# ---------------------------------------------------------------------------
_BRAND_PRIMARY = "#6f1d78"
_BRAND_LIGHT = "#d182d1"
_BRAND_BG_ALT = "#f7f0f8"
_BRAND_ORANGE = "#f57c20"

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
_STATIC_ROOT = Path(settings.BASE_DIR) / "static"
_FONT_DIR = _STATIC_ROOT / "fonts"

_FONTS_REGISTERED = False


def _get_logo_image():
    """Return a path (str) or ImageReader usable by reportlab drawImage.

    Supports both local filesystem and remote (S3/R2) storage backends.
    Returns *None* when no logo is configured or accessible.
    """
    try:
        from apps.content.models import PlatformLogoBlock

        block = PlatformLogoBlock.objects.filter(
            key="topbar_brand_logo", is_active=True,
        ).only("media_file").first()
        if not block or not block.media_file:
            return None

        # 1) Local filesystem (dev / non-S3)
        try:
            local_path = block.media_file.path
            if Path(local_path).exists():
                return local_path
        except NotImplementedError:
            pass  # S3 storage raises NotImplementedError for .path

        # 2) Remote storage — read into memory
        from reportlab.lib.utils import ImageReader

        block.media_file.open("rb")
        data = BytesIO(block.media_file.read())
        block.media_file.close()
        return ImageReader(data)
    except Exception:
        return None


def _get_platform_branding() -> dict:
    """Return platform name and website URL from admin settings."""
    try:
        from apps.content.services import public_branding_payload
        from apps.content.models import SiteLinks

        brand = public_branding_payload()
        links = SiteLinks.load()
        url = (links.website_url or "").replace("https://", "").replace("http://", "").rstrip("/")
        return {
            "name": brand.get("topbar_title") or "نوافــذ",
            "url": url,
        }
    except Exception:
        return {"name": "نوافــذ", "url": ""}


def _as_text(value) -> str:
    if value is None:
        return ""
    return str(value)


def _shape_ar(text: str) -> str:
    value = _as_text(text)
    try:
        import arabic_reshaper
        from bidi.algorithm import get_display

        reshaped = arabic_reshaper.reshape(value)
        # Cairo font is missing Arabic Presentation Forms B *isolated* glyphs.
        # Map them back to base Arabic characters so they remain visible.
        reshaped = reshaped.translate(_ISOLATED_TO_BASE)
        return get_display(reshaped)
    except Exception:
        return value


# Map Arabic Presentation Forms B isolated forms → base Arabic characters.
# Cairo (and many modern Arabic TTF fonts) only ship initial/medial/final
# presentation forms; isolated forms are expected to come from the base range.
_ISOLATED_TO_BASE = str.maketrans({
    0xFE80: 0x0621,  # HAMZA
    0xFE81: 0x0622,  # ALEF WITH MADDA ABOVE
    0xFE83: 0x0623,  # ALEF WITH HAMZA ABOVE
    0xFE85: 0x0624,  # WAW WITH HAMZA ABOVE
    0xFE87: 0x0625,  # ALEF WITH HAMZA BELOW
    0xFE89: 0x0626,  # YEH WITH HAMZA ABOVE
    0xFE8D: 0x0627,  # ALEF
    0xFE8F: 0x0628,  # BEH
    0xFE93: 0x0629,  # TEH MARBUTA
    0xFE95: 0x062A,  # TEH
    0xFE99: 0x062B,  # THEH
    0xFE9D: 0x062C,  # JEEM
    0xFEA1: 0x062D,  # HAH
    0xFEA5: 0x062E,  # KHAH
    0xFEA9: 0x062F,  # DAL
    0xFEAB: 0x0630,  # THAL
    0xFEAD: 0x0631,  # REH
    0xFEAF: 0x0632,  # ZAIN
    0xFEB1: 0x0633,  # SEEN
    0xFEB5: 0x0634,  # SHEEN
    0xFEB9: 0x0635,  # SAD
    0xFEBD: 0x0636,  # DAD
    0xFEC1: 0x0637,  # TAH
    0xFEC5: 0x0638,  # ZAH
    0xFEC9: 0x0639,  # AIN
    0xFECD: 0x063A,  # GHAIN
    0xFED1: 0x0641,  # FEH
    0xFED5: 0x0642,  # QAF
    0xFED9: 0x0643,  # KAF
    0xFEDD: 0x0644,  # LAM
    0xFEE1: 0x0645,  # MEEM
    0xFEE5: 0x0646,  # NOON
    0xFEE9: 0x0647,  # HEH
    0xFEED: 0x0648,  # WAW
    0xFEEF: 0x0649,  # ALEF MAKSURA
    0xFEF1: 0x064A,  # YEH
})


def _register_arabic_fonts():
    global _FONTS_REGISTERED
    if _FONTS_REGISTERED:
        return
    try:
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont

        regular = _FONT_DIR / "Cairo-Regular.ttf"
        bold = _FONT_DIR / "Cairo-Bold.ttf"
        if regular.exists():
            pdfmetrics.registerFont(TTFont("Cairo", str(regular)))
        if bold.exists():
            pdfmetrics.registerFont(TTFont("Cairo-Bold", str(bold)))
        _FONTS_REGISTERED = True
    except Exception:
        pass


def _get_font_name(bold: bool = False) -> str:
    _register_arabic_fonts()
    if _FONTS_REGISTERED:
        return "Cairo-Bold" if bold else "Cairo"
    return "Helvetica-Bold" if bold else "Helvetica"


# ---------------------------------------------------------------------------
# XLSX
# ---------------------------------------------------------------------------

def xlsx_response(filename: str, sheet_name: str, headers: list[str], rows: list[list]) -> HttpResponse:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = (sheet_name or "sheet")[:31]
    ws.sheet_view.rightToLeft = True

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="6F1D78", end_color="6F1D78", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.append([_as_text(h) for h in headers])
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for row in rows:
        ws.append([_as_text(v) for v in row])

    for row_cells in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row_cells:
            cell.alignment = cell_align

    for idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(idx)
        max_len = max(
            len(_as_text(ws.cell(row=r, column=idx).value)) for r in range(1, min(ws.max_row + 1, 50))
        )
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 40)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


# ---------------------------------------------------------------------------
# PDF – branded with logo, Arabic font, header/footer
# ---------------------------------------------------------------------------

def _build_header_footer_func(title: str, page_size):
    from reportlab.lib import colors
    from reportlab.lib.units import mm

    logo_image = _get_logo_image()
    branding = _get_platform_branding()
    font = _get_font_name(bold=True)
    font_regular = _get_font_name(bold=False)
    page_w = page_size[0]
    page_h = page_size[1]

    def _on_page(canvas, doc):
        canvas.saveState()

        # --- Header band ---
        header_h = 22 * mm
        canvas.setFillColor(colors.HexColor(_BRAND_PRIMARY))
        canvas.rect(0, page_h - header_h, page_w, header_h, fill=1, stroke=0)

        # Logo (right side for RTL)
        if logo_image:
            try:
                logo_w = 18 * mm
                logo_h = 18 * mm
                canvas.drawImage(
                    logo_image,
                    page_w - 12 * mm - logo_w,
                    page_h - header_h + 2 * mm,
                    width=logo_w,
                    height=logo_h,
                    preserveAspectRatio=True,
                    mask="auto",
                )
            except Exception:
                pass

        # Title text (right-aligned, next to logo)
        canvas.setFillColor(colors.white)
        canvas.setFont(font, 13)
        shaped_title = _shape_ar(title)
        canvas.drawRightString(
            page_w - 12 * mm - 22 * mm,
            page_h - header_h + 8 * mm,
            shaped_title,
        )

        # Platform name on left (from admin settings)
        platform_name = _shape_ar(branding["name"])
        canvas.setFont(font, 10)
        canvas.setFillColor(colors.HexColor(_BRAND_ORANGE))
        canvas.drawString(12 * mm, page_h - header_h + 9 * mm, platform_name)

        # Orange accent line under header
        canvas.setStrokeColor(colors.HexColor(_BRAND_ORANGE))
        canvas.setLineWidth(1.5)
        canvas.line(0, page_h - header_h, page_w, page_h - header_h)

        # --- Footer ---
        footer_y = 8 * mm
        canvas.setFont(font_regular, 7)
        canvas.setFillColor(colors.HexColor("#999999"))

        # Page number (centre)
        page_num = canvas.getPageNumber()
        canvas.drawCentredString(page_w / 2, footer_y, f"{page_num}")

        # Platform URL right (from admin settings)
        if branding["url"]:
            canvas.drawRightString(page_w - 12 * mm, footer_y, branding["url"])

        # Thin line above footer
        canvas.setStrokeColor(colors.HexColor("#e0e0e0"))
        canvas.setLineWidth(0.5)
        canvas.line(12 * mm, footer_y + 4 * mm, page_w - 12 * mm, footer_y + 4 * mm)

        canvas.restoreState()

    return _on_page


def pdf_response(
    filename: str,
    title: str,
    headers: list[str],
    rows: list[list],
    *,
    landscape: bool = False,
) -> HttpResponse:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Spacer, Table, TableStyle

    try:
        from reportlab.lib.pagesizes import landscape as rl_landscape
        page_size = rl_landscape(A4) if landscape else A4
    except Exception:
        page_size = A4

    _register_arabic_fonts()
    font = _get_font_name(bold=False)
    font_bold = _get_font_name(bold=True)

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=page_size,
        rightMargin=10 * mm,
        leftMargin=10 * mm,
        topMargin=28 * mm,
        bottomMargin=18 * mm,
    )

    content: list = [Spacer(1, 4 * mm)]

    table_data = [[_shape_ar(h) for h in headers]]
    for row in rows:
        table_data.append([_shape_ar(_as_text(v)) for v in row])

    table = Table(table_data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(_BRAND_PRIMARY)),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), font_bold),
                ("FONTSIZE", (0, 0), (-1, 0), 9),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
                ("TOPPADDING", (0, 0), (-1, 0), 6),
                ("FONTNAME", (0, 1), (-1, -1), font),
                ("FONTSIZE", (0, 1), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor(_BRAND_LIGHT)),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor(_BRAND_BG_ALT)]),
                ("TOPPADDING", (0, 1), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 1), (-1, -1), 4),
            ]
        )
    )
    content.append(table)

    on_page = _build_header_footer_func(title, page_size)
    doc.build(content, onFirstPage=on_page, onLaterPages=on_page)
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response

